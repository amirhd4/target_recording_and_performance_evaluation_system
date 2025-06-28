import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import jdatetime
from decimal import Decimal, ROUND_DOWN
import io
from django.http import HttpResponse
from django.db.models import Count, Sum, F, Value, OuterRef, Subquery, Case, When, DecimalField
from django.db.models.functions import Coalesce
from accounts.models import CustomUser, MonthlyScore 
from .models import BoardAdjustment
from django.utils.encoding import escape_uri_path
import datetime
from jdatetime import datetime as jdt_datetime


def get_jalali_date():
    """Returns current Jalali year, month, day as a tuple."""
    today_gregorian = datetime.date.today()
    today_jalali = jdatetime.date.fromgregorian(date=today_gregorian)
    return (today_jalali.year, today_jalali.month, today_jalali.day)


def get_jalali_month_name(month_number):
    """Returns Jalali month name for a given month number."""
    month_names = [
        "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
        "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
    ]
    if 1 <= month_number <= 12:
        # TODO: This is for past month
        return month_names[month_number - 2]
    return str(month_number)


def monthly_scores_summary_excel_func(request):
    """
    Generates an Excel file of the monthly score summary, including board adjustments.
    Uses openpyxl for generation and styling.
    """

    # today_jalali_parts = get_jalali_date()
    try:
        jalali_year = int(request.GET.get("year", jdt_datetime.now().year))
        jalali_month = int(request.GET.get("month", jdt_datetime.now().month))
    except Exception as e:
        jalali_year = 0
        jalali_month = 0
        print(e)

    # jalali_year = today_jalali_parts[0]
    # jalali_month = today_jalali_parts[1]
    jalali_month_name = get_jalali_month_name(jalali_month)


    # --- واکشی و تجمع داده‌ها (مشابه منطق خلاصه ماهانه در ViewSet) ---

    # فیلتر امتیازات ماهانه بر اساس سال و ماه، و حذف امتیازاتی که scorer ندارند
    month_queryset = MonthlyScore.objects.filter(
        year=jalali_year,
        month=jalali_month,
        scorer__isnull=False
    )

    # اگر هیچ امتیاز ماهانه‌ای برای ماه جاری پیدا نشد، پاسخ مناسب برگردانید
    if not month_queryset.exists():
         response = HttpResponse("داده امتیاز ماهانه برای ماه جاری یافت نشد.", status=404)
         response['Content-Type'] = 'text/plain; charset=utf-8' # اطمینان از نمایش صحیح پیام
         return response

    # محاسبه مقدار وزن‌دهی شده امتیاز (مشابه ViewSet)
    # ابتدا تعداد کاربران در هر سطح دسترسی را برای وزن‌دهی واکشی کنید
    user_counts_by_level_qs = CustomUser.objects.filter(access_level__gt=-2) \
                                                .values('access_level') \
                                                .annotate(count=Count('id'))
    access_level_counts = {item['access_level']: item['count'] for item in user_counts_by_level_qs}
    access_level_counts_decimal = {
        level: Decimal(count) for level, count in access_level_counts.items()
    }

    # تعریف آنوتیشن برای محاسبه امتیاز وزن‌دهی شده
    weighted_value_annotation = Case(
        When(scorer__access_level__isnull=True, then=Value(Decimal(0))),
        When(scorer__access_level=1, then=F('value') / Value(access_level_counts_decimal.get(1, Decimal(1)))),
        When(scorer__access_level=2, then=F('value') / Value(access_level_counts_decimal.get(2, Decimal(1)))),
        When(scorer__access_level=3, then=F('value') / Value(access_level_counts_decimal.get(3, Decimal(1)))),
        When(scorer__access_level=0, then=F('value') / Value(access_level_counts_decimal.get(0, Decimal(1)))),
        When(scorer__access_level=-1, then=F('value') / Value(access_level_counts_decimal.get(-1, Decimal(1)))),
        default=F('value'), # برای سایر سطوح دسترسی احتمالی
        output_field=DecimalField()
    )

    # تجمع امتیازات وزن‌دهی شده بر اساس کاربر هدف و محاسبه مجموع امتیاز ماهانه محاسبه شده (به ازای هر کاربر)
    aggregated_scores_by_user_qs = month_queryset.values(
        'target__id',
        'target__first_name',
        'target__last_name',
        'target__username',
        'target__access_level',
    ).annotate(
        total_monthly_score=Coalesce(Sum(weighted_value_annotation), Value(Decimal(0)))
    ) # این QuerySet شامل total_monthly_score به ازای هر کاربر است

    # واکشی امتیازات تنظیم شده هیئت مدیره با استفاده از Subquery و آنوتیشن روی QuerySet بالا
    board_adjustment_subquery = BoardAdjustment.objects.filter(
        user=OuterRef('target__id') # اتصال به کاربر هدف در تجمع
    ).values('adjustment_value')[:1] # فقط آخرین مقدار تنظیم شده را بگیرید

    aggregated_scores_with_adjustment_qs = aggregated_scores_by_user_qs.annotate(
        board_adjustment=Coalesce(
            Subquery(board_adjustment_subquery, output_field=DecimalField()),
            Value(Decimal(0)) # اگر تنظیمی برای کاربر نبود، صفر در نظر بگیرید
        )
    )

    # واکشی داده ها از دیتابیس (اینجا داده ها شامل total_monthly_score و board_adjustment به ازای هر کاربر است)
    aggregated_scores_with_adjustment_list = list(aggregated_scores_with_adjustment_qs)

    # --- آماده سازی لیست داده برای اکسل و محاسبه مجموع نهایی با تنظیم در پایتون و اضافه کردن نام سطح دسترسی ---
    excel_data = []
    access_level_names_map_full = dict(CustomUser.LEVEL_CHOICES) # استفاده از Mapping کامل سطوح دسترسی

    # آماده سازی داده‌های اصلی شیت کاربران
    
    

    for entry in aggregated_scores_with_adjustment_list:
        total_score = entry['total_monthly_score']
        adjustment = entry['board_adjustment']
        adjusted_total = total_score + adjustment
        access_level_display = access_level_names_map_full.get(entry['target__access_level'], f'سطح {entry['target__access_level']}')


        excel_data.append({
            'user_id': entry['target__id'],
            'first_name': entry['target__first_name'],
            'last_name': entry['target__last_name'],
            'username': entry['target__username'],
            'access_level': entry['target__access_level'], # نگهداری عدد سطح دسترسی برای مرتب سازی اگر نیاز است
            'access_level_name': access_level_display, # نام قابل نمایش سطح دسترسی
            'total_monthly_score': total_score.quantize(Decimal('0.001'), rounding=ROUND_DOWN), # نمایش با 3 رقم اعشار
            'board_adjustment': adjustment.quantize(Decimal('0.000'), rounding=ROUND_DOWN), # نمایش با 3 رقم اعشار
            'total_monthly_score_adjusted': adjusted_total.quantize(Decimal('0.001'), rounding=ROUND_DOWN), # نمایش با 3 رقم اعشار
        })

    # مرتب سازی نهایی داده ها (مثال: بر اساس سطح دسترسی و سپس مجموع نهایی با تنظیم، نزولی)
    # شما می توانید این مرتب سازی را بر اساس نیاز خود تغییر دهید
    # اینجا از کلید 'access_level' که در دیکشنری های excel_data وجود دارد، استفاده می شود.
    excel_data = sorted(excel_data, key=lambda x: (x['total_monthly_score_adjusted']), reverse=True)

    # --- محاسبه آمار کلی سطوح در پایتون (برای شیت آمار کلی سطوح) ---
    # این بخش با پیمایش روی داده های کاربران که total_monthly_score محاسبه شده دارند، انجام می شود
    # تا از خطای تجمع روی تجمع جلوگیری شود.
    overall_stats_dict = {} # دیکشنری برای نگهداری جمع امتیاز و تعداد کاربران هر سطح

    # پیمایش روی داده های کاربران که شامل total_monthly_score به ازای هر کاربر است
    for entry in aggregated_scores_by_user_qs: # استفاده از QuerySet قبل از آنوتیشن board_adjustment
        level = entry['target__access_level']
        total_score = entry['total_monthly_score'] # امتیاز محاسبه شده به ازای هر کاربر

        if level not in overall_stats_dict:
            level_name = access_level_names_map_full.get(level, f'سطح {level}')
            overall_stats_dict[level] = {
                 'level': level, # <--- اضافه کردن کلید 'level'
                 'level_name': level_name,
                 'count_with_score': 0, # تعداد کاربرانی که در این ماه امتیاز دریافت کرده اند در این سطح
                 'total_score_sum': Decimal(0), # مجموع امتیاز محاسبه شده این کاربران در این سطح
            }

        overall_stats_dict[level]['count_with_score'] += 1 # افزایش تعداد کاربران با امتیاز
        overall_stats_dict[level]['total_score_sum'] += total_score # جمع کردن امتیاز محاسبه شده این کاربران در این سطح


    # تبدیل دیکشنری آمار کلی به لیست برای اکسل
    overall_excel_data = []
    # واکشی تعداد کل کاربران در هر سطح (برای محاسبه 'امتیاز نداده اند')
    user_total_counts_by_level_qs = CustomUser.objects.filter(access_level__gt=-2).values('access_level').annotate(count=Count('id'))
    user_total_counts_dict = {item['access_level']: item['count'] for item in user_total_counts_by_level_qs}

    for level, stats in overall_stats_dict.items():
        total_users_in_level = user_total_counts_dict.get(level, 0)
        no_givers_count = total_users_in_level - stats['count_with_score'] # تعداد کاربرانی که امتیازی دریافت نکرده اند

        overall_excel_data.append({
            'level': stats['level'], # <--- استفاده از کلید 'level'
            'level_name': stats['level_name'],
            'count_with_score': stats['count_with_score'],
            'no_givers_count': no_givers_count,
            'total_score_sum': stats['total_score_sum'].quantize(Decimal('0.001'), rounding=ROUND_DOWN), # نمایش با 3 رقم اعشار
        })

    # مرتب سازی داده آمار کلی (مثال: بر اساس سطح دسترسی عددی)
    # اینجا از کلید 'level' که در دیکشنری های overall_excel_data اضافه شده، استفاده می شود.
    overall_excel_data = sorted(overall_excel_data, key=lambda x: x['level'])


    # --- واکشی کاربران بدون امتیازدهی (برای شیت کاربران بدون امتیازدهی) ---
    # پیدا کردن scorerهایی که در ماه جاری امتیاز داده اند
    scorers_in_month_ids = month_queryset.values_list('scorer__id', flat=True).distinct()
    # پیدا کردن کاربرانی در سطوح 1, 2, 3 که در لیست scorers_in_month_ids نیستند و باید امتیاز می دادند
    no_score_givers_qs = CustomUser.objects.filter(
        access_level__in=[1, 2, 3] # فرض بر اینکه این سطوح باید امتیاز دهند
    ).exclude(
        id__in=scorers_in_month_ids
    ).order_by('access_level', 'first_name', 'last_name') # مرتب سازی

    no_score_givers_list_data = list(no_score_givers_qs)


    # --- تولید فایل اکسل با openpyxl و استایل دهی ---

    # ایجاد Workbook و Worksheet اصلی (شیت خلاصه کاربران)
    workbook = openpyxl.Workbook()
    ws_users = workbook.active # استفاده از شیت فعال پیش فرض
    ws_users.title = f'خلاصه ماه {jalali_month_name} {jalali_year}' # عنوان شیت

    # راست چین کردن شیت
    ws_users.sheet_view.rightToLeft = True

    # تعریف استایل ها
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid") # خاکستری روشن

    score_number_format = '0.000' # فرمت اعداد اعشاری (3 رقم بعد از اعشار)
    adjusted_total_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # سبز روشن


    # اضافه کردن عنوان و زیرعنوان در بالای شیت اصلی (ردیف های 1 و 2)
    title = f'خلاصه امتیازات ماهانه {jalali_month_name} از سال {jalali_year}'
    subtitle = 'کارت اعتبار فارس'
    
    # Merge سلول ها برای عنوان و زیرعنوان (جدول اصلی 7 ستون دارد)
    ws_users.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws_users['A1'].value = title
    ws_users['A1'].font = Font(bold=True, size=14)
    ws_users['A1'].alignment = center_align

    ws_users.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    ws_users['A2'].value = subtitle
    ws_users['A2'].font = Font(size=12)
    ws_users['A2'].alignment = center_align

    # یک ردیف خالی برای فاصله (ردیف 3)


    # نوشتن سربرگ جدول کاربران (شروع از ردیف 4)
    headers_users = ["ردیف", "نام", "نام کاربری", "سطح دسترسی", "امتیاز محاسبه شده ماهانه", "تنظیم هیئت مدیره", "مجموع نهایی امتیاز"]
    
    # نوشتن سربرگ ها و اعمال استایل سربرگ
    for col_idx, header_text in enumerate(headers_users):
         cell = ws_users.cell(row=4, column=col_idx + 1, value=header_text)
         cell.font = header_font
         cell.alignment = center_align
         cell.border = thin_border_style
         cell.fill = header_fill


    # نوشتن ردیف های داده کاربران (شروع از ردیف 5)
    row_num = 5
    # ترتیب کلیدهای داده برای نوشتن در ردیف ها (مطابق سربرگ ها بعد از ستون ردیف)
    data_keys_order_users = ['first_name', 'last_name', 'username', 'access_level_name', 'total_monthly_score', 'board_adjustment', 'total_monthly_score_adjusted']


    for index, data_row in enumerate(excel_data):
        # نوشتن شماره ردیف (ستون ۱)
        ws_users.cell(row=row_num, column=1, value=index + 1).alignment = center_align # شماره ردیف وسط چین

        # نوشتن بقیه داده ها و اعمال استایل به هر سلول
        # ستون ۲: نام کامل (ترکیب اول و آخر)
        cell = ws_users.cell(row=row_num, column=2, value=f"{data_row.get('first_name', '')} {data_row.get('last_name', '')}".strip())
        cell.alignment = center_align
        cell.border = thin_border_style

        # ستون ۳: نام کاربری
        cell = ws_users.cell(row=row_num, column=3, value=data_row.get('username', ''))
        cell.alignment = center_align
        cell.border = thin_border_style

        # ستون ۴: سطح دسترسی
        cell = ws_users.cell(row=row_num, column=4, value=data_row.get('access_level_name', ''))
        cell.alignment = center_align
        cell.border = thin_border_style

        # ستون ۵: امتیاز محاسبه شده ماهانه
        cell = ws_users.cell(row=row_num, column=5, value=data_row.get('total_monthly_score', Decimal(0)))
        cell.alignment = center_align
        cell.border = thin_border_style
        cell.number_format = score_number_format

        # ستون ۶: تنظیم هیئت مدیره
        cell = ws_users.cell(row=row_num, column=6, value=data_row.get('board_adjustment', Decimal(0)))
        cell.alignment = center_align
        cell.border = thin_border_style
        cell.number_format = score_number_format

        # ستون ۷: مجموع نهایی امتیاز
        cell = ws_users.cell(row=row_num, column=7, value=data_row.get('total_monthly_score_adjusted', Decimal(0)))
        cell.alignment = center_align
        cell.border = thin_border_style
        cell.number_format = score_number_format
        cell.font = Font(bold=True) # بولد کردن
        cell.fill = adjusted_total_fill # پس‌زمینه

        row_num += 1 # رفتن به ردیف بعدی


    # تنظیم عرض ستون ها برای شیت کاربران
    for col_idx in range(1, ws_users.max_column + 1):
         max_length = 0
         column = get_column_letter(col_idx)
         # بررسی محتوا در ردیف های هدر، عنوان و داده (از ردیف 1 تا آخر)
         for row_idx in range(1, ws_users.max_row + 1):
              cell = ws_users[column + str(row_idx)]
              try:
                   if cell.value is not None:
                        cell_value_str = str(cell.value)
                        # بررسی فرمت عددی برای تخمین طول بهتر
                        if cell.number_format and cell.number_format != 'General':
                             if isinstance(cell.value, (int, float, Decimal)):
                                  if cell.number_format == '0.000': cell_value_str = f"{cell.value:.3f}" # مثال برای فرمت اعشاری
                                  else: cell_value_str = str(cell.value)
                             else: cell_value_str = str(cell_value_str)
                        else: cell_value_str = str(cell_value_str) # برای متن یا سایر موارد


                        max_length = max(max_length, len(cell_value_str))
              except Exception as e:
                   pass # نادیده گرفتن خطاها در حین محاسبه طول


         # تنظیم عرض ستون با کمی فضای اضافه
         adjusted_width = (max_length + 2) * 0.9 # ضریب 0.9 معمولاً برای فونت های پیش فرض اکسل مناسب تر است
         # اعمال محدودیت برای حداقل و حداکثر عرض
         if adjusted_width > 60: adjusted_width = 60
         elif adjusted_width < 10: adjusted_width = 10

         ws_users.column_dimensions[column].width = adjusted_width


    # --- نوشتن داده در شیت "آمار کلی سطوح" ---
    # ایجاد شیت جدید
    ws_overall = workbook.create_sheet(title="آمار کلی سطوح")
    ws_overall.sheet_view.rightToLeft = True # راست چین کردن

    # نوشتن سربرگ ها در شیت آمار کلی (شروع از ردیف 1)
    overall_headers = ["سطح دسترسی", "تعداد کاربران (با امتیاز)", "تعداد کاربران (بدون امتیاز)", "مجموع امتیاز محاسبه شده کل سطح"]
    ws_overall.append(overall_headers)
    # اعمال استایل سربرگ ها
    for col_idx, cell in enumerate(ws_overall[1]): # ردیف 1 (index 0)
         cell.font = header_font; cell.alignment = center_align; cell.border = thin_border_style; cell.fill = header_fill

    # نوشتن داده ها در شیت آمار کلی (شروع از ردیف 2)
    for row_data in overall_excel_data: # استفاده از overall_excel_data که در پایتون آماده شده
         ws_overall.append([
             row_data['level_name'],
             row_data['count_with_score'],
             row_data['no_givers_count'],
             row_data['total_score_sum'],
         ])
         # اعمال استایل به سلول های داده
         current_row = ws_overall[ws_overall.max_row]
         for col_idx, cell in enumerate(current_row):
              cell.alignment = center_align; cell.border = thin_border_style
              # اعمال فرمت عدد برای ستون مجموع امتیاز (ستون 4)
              if col_idx == 3: # ستون چهارم (index 3)
                   cell.number_format = score_number_format


    # تنظیم عرض ستون ها برای شیت آمار کلی
    for col_idx in range(1, ws_overall.max_column + 1):
         max_length = 0
         column = get_column_letter(col_idx)
         for row_idx in range(1, ws_overall.max_row + 1): # بررسی از ردیف 1 تا آخر
              cell = ws_overall[column + str(row_idx)]
              try:
                   if cell.value is not None:
                        cell_value_str = str(cell.value)
                        if cell.number_format and cell.number_format != 'General':
                             if isinstance(cell.value, (int, float, Decimal)):
                                  if cell.number_format == '0.000': cell_value_str = f"{cell.value:.3f}"
                                  else: cell_value_str = str(cell.value)
                             else: cell_value_str = str(cell_value_str)
                        else: cell_value_str = str(cell_value_str)
                        max_length = max(max_length, len(cell_value_str))
              except Exception as e: pass

         adjusted_width = (max_length + 2) * 0.9
         if adjusted_width > 60: adjusted_width = 60
         elif adjusted_width < 10: adjusted_width = 10
         ws_overall.column_dimensions[column].width = adjusted_width


    # --- نوشتن داده در شیت "کاربران بدون امتیازدهی" ---
    # ایجاد شیت جدید
    ws_no_givers_list = workbook.create_sheet(title="کاربران بدون امتیازدهی")
    ws_no_givers_list.sheet_view.rightToLeft = True # راست چین کردن

    # نوشتن سربرگ ها در شیت کاربران بدون امتیازدهی (شروع از ردیف 1)
    no_givers_list_headers = ["ردیف", "نام", "نام کاربری", "سطح دسترسی"]
    ws_no_givers_list.append(no_givers_list_headers)
    # اعمال استایل سربرگ ها
    for col_idx, cell in enumerate(ws_no_givers_list[1]): # ردیف 1 (index 0)
         cell.font = header_font; cell.alignment = center_align; cell.border = thin_border_style; cell.fill = header_fill

    # نوشتن داده ها در شیت کاربران بدون امتیازدهی (شروع از ردیف 2)
    total_no_givers_counter = 0 # شمارنده ردیف
    for user in no_score_givers_list_data: # استفاده از no_score_givers_list_data که در پایتون آماده شده
         total_no_givers_counter += 1 # افزایش شمارنده ردیف
         user_level_name = access_level_names_map_full.get(user.access_level, f"سطح {user.access_level}")
         ws_no_givers_list.append([
             total_no_givers_counter, # شماره ردیف
             f"{user.first_name} {user.last_name}".strip(), # نام کامل
             user.username, # نام کاربری
             user_level_name, # نام سطح دسترسی
         ])
         # اعمال استایل به سلول های داده
         current_row = ws_no_givers_list[ws_no_givers_list.max_row]
         for cell in current_row:
              cell.alignment = center_align; cell.border = thin_border_style


    # تنظیم عرض ستون ها برای شیت کاربران بدون امتیازدهی
    for col_idx in range(1, ws_no_givers_list.max_column + 1):
         max_length = 0
         column = get_column_letter(col_idx)
         for row_idx in range(1, ws_no_givers_list.max_row + 1): # بررسی از ردیف 1 تا آخر
              cell = ws_no_givers_list[column + str(row_idx)]
              try:
                   if cell.value is not None:
                        cell_value_str = str(cell.value)
                        if cell.number_format and cell.number_format != 'General':
                             if isinstance(cell.value, (int, float, Decimal)):
                                  if cell.number_format == '0.000': cell_value_str = f"{cell.value:.3f}"
                                  else: cell_value_str = str(cell.value)
                             else: cell_value_str = str(cell_value_str)
                        else: cell_value_str = str(cell_value_str)
                        max_length = max(max_length, len(cell_value_str))
              except Exception as e: pass

         adjusted_width = (max_length + 2) * 0.9
         if adjusted_width > 60: adjusted_width = 60
         elif adjusted_width < 10: adjusted_width = 10
         ws_no_givers_list.column_dimensions[column].width = adjusted_width


    # 8. ذخیره Workbook در حافظه و آماده سازی پاسخ HTTP
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    # تنظیم نام فایل با تاریخ شمسی و استفاده از escape_uri_path
    file_name = f"خلاصه-امتیازات-ماهانه-{jalali_year}-{jalali_month}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{escape_uri_path(file_name)}"'

    return response